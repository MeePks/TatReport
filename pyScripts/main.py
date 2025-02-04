from configparser import ConfigParser
import functions as fn
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta
from sqlalchemy import text
import sendEMail as se

# Get the start date (first day of the previous month)
first_day_of_current_month = datetime.now().replace(day=1)
start_date = (first_day_of_current_month - timedelta(days=1)).replace(day=1)

# Get the end date (last day of the previous month)
end_date = first_day_of_current_month - timedelta(days=1)
output_file = rf'Y:\Data\Retail\WalmartMX\Development\Pikesh.Maharjan\RetailReports\TAT\Tat_ '+rf'{start_date}_{end_date}.xlsx'

# Convert to `date` objects if needed
start_date = start_date.date()
end_date = end_date.date()

def adjust_column_width(worksheet):
    """
    Adjust column widths based on the content in both data and header.
    """
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get the column letter
        # Check the header (first row)
        if col[0].value:
            max_length = len(str(col[0].value))
        
        # Check the data (rest of the rows)
        for cell in col:
            try:
                if cell.value:  # Check if the cell has a value
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        
        # Adjust the column width
        adjusted_width = max_length + 5  # Add some padding
        worksheet.column_dimensions[col_letter].width = adjusted_width

#defining functions to get weigheted mean and Median
# Define a function to calculate weighted mean
def weighted_mean(group):
    return round((group["AvgTAT(Mean)"] * group["NoOfFiles"]).sum() / group["NoOfFiles"].sum(), 2)

# Define a function to calculate weighted median
def weighted_median(group):
    weighted_values = group.loc[group.index.repeat(group["NoOfFiles"])]
    return round(weighted_values["TATMedian"].median(), 2)

#reading configuration file
config=ConfigParser()
config.read('.\pyScripts\config.ini')
centralized_server=config['sqlconn']['centralized_server']
centralized_db=config['sqlconn']['centralized_db']
sql_script='ReportGeneration.sql'
perofOLA='TatPerReport.sql'

#connecting to the centralized server in which there is list of all the servers

centralized_conn=fn.open_alchemy_conn(centralized_server,centralized_db)
df_server_list=pd.read_sql_table('SSIS_ConfigurationInfo',centralized_conn)
df_Ola_details=pd.read_sql_table('___RetailOLA',centralized_conn)


#query to check if a stored proc exists or not
query_proc_exists=r"SELECT COUNT(*) AS ProcExist FROM INFORMATION_SCHEMA.ROUTINES WHERE ROUTINE_NAME = '__getTat' AND ROUTINE_TYPE = 'PROCEDURE'"

# Initialize an empty list to store the results
results = []
df_results=pd.DataFrame()
df_tats=pd.DataFrame()
df_olaTat_details=pd.DataFrame()

#open sql script to read the query
with open(sql_script,'r') as file:
    report_generate_query=file.read()

with open(perofOLA,'r') as file:
    ola_generate_query=file.read()

sp_exec_query=f'''
EXEC [dbo].[__getTat] @startDate = '{start_date}', @EndDate = '{end_date}'
'''

#iterating through server list
for index,rows in df_server_list.iterrows():
    #iterate only through active audits
    if rows['ActiveFlag']:
       audit_conn=fn.open_alchemy_conn(rows['SourceServerName'],rows['InventoryLogDB']) #opening connection in audits
       df_proc=pd.read_sql_query(query_proc_exists,audit_conn,index_col=None) #checking if proc exists
       #Getting report with the help of sql script Reportgeneration.sql
       try:
            with audit_conn.connect() as connection:
                connection.execute(text(sp_exec_query.strip()))
            df_Ola_details.to_sql('___OLADetails',audit_conn,schema='dbo',if_exists='replace')

            #getting Tatsummary report for each audits
            df_result=pd.read_sql_query(report_generate_query,audit_conn,index_col=None)
            df_result['AuditName']=rows['AuditName']
            if not df_result.empty:
                df_results=pd.concat([df_results,df_result])

            #getting percentage of Tat fullfilled
                

            df_olaTat_detail=pd.read_sql_query(ola_generate_query,audit_conn,index_col=None)
            df_olaTat_detail['AuditName']=rows['AuditName']
            if not df_olaTat_detail.empty:
                df_olaTat_details=pd.concat([df_olaTat_details,df_olaTat_detail])

            #getting entire tat details so to produce tat details of not getting fullfillled
            df_tat=pd.read_sql_table('tat',audit_conn,schema='dbo',index_col=None)
            if not df_tat.empty:
                df_tats=pd.concat([df_tats,df_tat])
       except:
           print(f'Error executing script for {rows["AuditName"]}')

        #Appending the details of proc exists sql statements
       results.append({
           'AuditName':rows['AuditName'],
           'ServerName':rows['SourceServerName'],
           'DatabaseName':rows['InventoryLogDB'],
           'ProcExists':df_proc['ProcExist'][0]
           })
       
#getting details of tat and getting if it fullfilled the OLA
df_results_column=['AuditName','Frequency','NoOfFiles','AvgTAT(Mean)','TATMedian','OLAMetPer','OLANotMetPer']
df_olaTat_details=df_olaTat_details[df_results_column]
       
#converting list to dataframe for easy data manipulation      
df_proc_exists_list=pd.DataFrame(results)
df_proc_exists_list.to_csv('TatProcExistlist.csv',index=False)

#generating overall Report and changeing the order of columns
df_results_column=['AuditName','Frequency','NoOfFiles','AvgTAT(Mean)','TATMedian']
df_results=df_results[df_results_column]
#df_results.to_csv('TatReportDetails.csv', index=False)

# Group by Frequency and calculate weighted mean and median
weighted_results = (
    df_results.groupby("Frequency", group_keys=False)
    .apply(lambda group: pd.Series({
        "Weighted Mean TAT": weighted_mean(group),
        "Weighted Median TAT": weighted_median(group)
    }))
    .reset_index()
)
weighted_results.insert(0, "FromDate", start_date)
weighted_results.insert(1, "ToDate", end_date)

#pivoting the table
df_pivot=df_results.pivot_table(
    index='AuditName',
    columns='Frequency',
    values=['AvgTAT(Mean)','TATMedian'],
    aggfunc='first'
)

# Reorder the columns to have mean and median for each frequency together
df_pivot = df_pivot.reorder_levels([1, 0], axis=1).sort_index(axis=1, level=0)

# Flatten the MultiIndex columns (if any)
df_pivot.columns = [
    f"{freq} (Mean)" if metric == "AvgTAT(Mean)" else f"{freq} (Median)"
    for freq, metric in df_pivot.columns
]

# Reset index for a clean DataFrame
df_pivot_overall = df_pivot.reset_index()

# Display the pivoted DataFrame
#df_pivot_overall.to_csv('TatReport.csv',index=False)

#pivoting the table based on Mean
df_pivot=df_results.pivot_table(
    index='AuditName',
    columns='Frequency',
    values=['AvgTAT(Mean)'],
    aggfunc='first'
)

# Remove the multi-level column name (flatten it)
df_pivot.columns.name = None  # Remove the 'frequency' header

# Rename the columns to clean any extra metadata
df_pivot.columns = [col if isinstance(col, str) else col[1].strip() for col in df_pivot.columns]
df_pivot_mean = df_pivot.reset_index()  # Reset the index to keep AuditName as a column

# Display the pivoted DataFrame
#df_pivot_mean.to_csv('TatReport_Mean.csv',index=False)

#pivoting the table based on Median
df_pivot=df_results.pivot_table(
    index='AuditName',
    columns='Frequency',
    values='TATMedian',
    aggfunc='first'
)

# Remove the multi-level column name (flatten it)
df_pivot.columns.name = None  # Remove the 'frequency' header


# Rename the columns to clean any extra metadata
df_pivot.columns = [col if isinstance(col, str) else col[1].strip() for col in df_pivot.columns]
df_pivot_median = df_pivot.reset_index()  # Reset the index to keep AuditName as a column

# Display the pivoted DataFrame
#df_pivot_median.to_csv('TatReport_Median.csv',index=False)


# Filter data for each frequency
frequencies = ["Daily", "Weekly", "Monthly"]
charts_positions = {"Daily": "H5", "Weekly": "H20", "Monthly": "H35"}

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        weighted_results.to_excel(writer, sheet_name='AggReport', index=False)
        df_olaTat_details.to_excel(writer, sheet_name=f'TatDetails', index=False)
        df_pivot_overall.to_excel(writer, sheet_name="PivotedDetails", startrow=0, startcol=0, index=False)
        df_pivot_mean.to_excel(writer, sheet_name="MeanDetails", startrow=0, startcol=0, index=False)
        df_pivot_median.to_excel(writer, sheet_name="MedianDetails", startrow=0, startcol=0, index=False)
        df_Ola_details.to_excel(writer,sheet_name="OLA", startrow=0, startcol=0, index=False)

        # Access the workbook and worksheet objects
        workbook = writer.book

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            if sheet_name == "PivotedDetails":
                sheet.freeze_panes = sheet["B1"]
                

            # Adjust column widths for the current sheet
            adjust_column_width(sheet)


        

# Add tables and headers with openpyxl
wb = load_workbook(output_file)

# Function to add a table
def add_table(ws, startrow, startcol, df, table_name):
    endrow = startrow + len(df) + 1
    endcol = startcol + len(df.columns) -1
    table = Table(
        displayName=table_name,
        ref=f"{chr(65 + startcol)}{startrow + 1}:{chr(65 + endcol)}{endrow}"
    )
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

# Add tables to the first sheet (AggReport)
ws1 = wb["AggReport"]
add_table(ws1, 0, 0, weighted_results, "AggReportTable")

# Add tables to the second sheet (TatDetails)
ws2 = wb["TatDetails"]
add_table(ws2, 0, 0, df_olaTat_details, "TatDetailsTable")

# Add tables to the first sheet (AggReport)
ws3 = wb["PivotedDetails"]
#add_table(ws3, 0, 0, df_pivot_overall, "PivotedDetailsTable")

# Add tables to the second sheet (TatDetails)
ws4 = wb["MeanDetails"]
add_table(ws4, 0, 0, df_pivot_mean, "MeanDetailstable")

# Add tables to the second sheet (TatDetails)
ws5 = wb["MedianDetails"]
add_table(ws5, 0, 0, df_pivot_median, "MedianDetailsTable")

ws6 = wb["OLA"]
add_table(ws6, 0, 0, df_Ola_details, "OLA")


# Save the workbook
wb.save(output_file)

se.send_email(from_date, to_date, report_date, path, schedule)